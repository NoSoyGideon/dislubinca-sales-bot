# src/services/excel_service.py

import os
import shutil
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
from services.base_excel_service import BaseExcelService
import time
from config.excel_map_config import CobranzaDiariaMap as MapCD,ContactoMatutinoMap

class ExcelService(BaseExcelService):
    def __init__(self, dropbox_service):
        """
        Servicio definitivo de persistencia física en Excel.
        Maneja archivos .xlsm protegiendo el código VBA y las macros nativas.
        """
        self.dropbox = dropbox_service
        self.MAIN_FOLDER = "/CONTACTO MATUTINO/"
        # Forzar rutas absolutas relativas al directorio de este archivo de servicio
        self.tmp_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../tmp"))
        self.base_dir = os.path.join(self.tmp_dir, "plantillas_base")
        os.makedirs(self.tmp_dir, exist_ok=True)

    def _calcular_nombres(self, fecha_str: str):
        """Genera centralizadamente los nombres de archivos (.xlsm) y pestañas diarias."""
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        meses = {1:"ENE", 2:"FEB", 3:"MAR", 4:"ABR", 5:"MAY", 6:"JUN",
                 7:"JUL", 8:"AGO", 9:"SEP", 10:"OCT", 11:"NOV", 12:"DIC"}
        dias = {0:"LUNES", 1:"MARTES", 2:"MIERCOLES", 3:"JUEVES", 4:"VIERNES", 5:"SABADO", 6:"DOMINGO"}
        
        mes_tag = meses[dt.month]
        anio_tag = str(dt.year)[2:]
        
        archivo_matutino = f"CONTACTO MATUTINO {mes_tag}-{anio_tag}.xlsm"
        archivo_cobranza = f"REPORTE DIARIO DE COBRANZA {mes_tag}-{anio_tag}.xlsx"
        pestana_dia = f"{dias[dt.weekday()]} {dt.strftime('%d-%m')}"
        
        return archivo_matutino, archivo_cobranza, pestana_dia

    def _asegurar_pestana_desde_plantilla(self, wb: openpyxl.Workbook, nombre_pestana: str, fecha_str: str = None):
        """
        [MÉTODO CORE UNIFICADO FIX CASE-INSENSITIVE]
        Garantiza que exista una pestaña buscando el nombre sin importar mayúsculas/minúsculas.
        Si no existe, clona 'Plantilla' e inyecta la fecha en G4.
        """
        # 1. Buscamos si ya existe la pestaña (ignorando mayúsculas/minúsculas)
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().upper() == nombre_pestana.strip().upper():
                return wb[sheet_name]

        # 2. Si se trata de la mismísima pestaña 'PLANTILLA' y no estaba, retornamos la que haya
        if nombre_pestana.strip().upper() == "PLANTILLA":
            for sheet in wb.sheetnames:
                if sheet.strip().upper() == "PLANTILLA":
                    return wb[sheet]

        print(f"⚙️ [ExcelService] Clonando pestaña '{nombre_pestana}' desde 'Plantilla'...")
        
        # 3. Ubicamos la pestaña base 'Plantilla'
        pestana_plantilla = None
        for sheet in wb.sheetnames:
            if sheet.strip().upper() == "PLANTILLA":
                pestana_plantilla = wb[sheet]
                break

        if not pestana_plantilla:
            pestana_plantilla = wb.worksheets[0]

        # 4. Clonamos y asignamos título limpio
        nueva_ws = wb.copy_worksheet(pestana_plantilla)
        nueva_ws.title = nombre_pestana
        
        # Inyectamos la fecha en G4 solo si es pestaña diaria válida
        if fecha_str:
            try:
                dt = datetime.strptime(fecha_str, "%Y-%m-%d")
                nueva_ws["G4"] = dt.strftime("%d/%m/%Y")
            except ValueError:
                pass
        
        return nueva_ws
    def _garantizar_archivo_y_pestana(self, nombre_archivo: str, plantilla_base_nombre: str, pestana_target: str, ruta_local: str, fecha_str: str = None) -> openpyxl.Workbook:
        """
        [GARANTE DE ARCHIVO Y NUBE]
        Garantiza la existencia del libro en Dropbox/local. Delega la creación de pestañas
        a _asegurar_pestana_desde_plantilla.
        """
        ruta_dropbox = f"{self.MAIN_FOLDER}{nombre_archivo}"
        
        # 1. Intentar descargar desde la nube de Dropbox si no está local
        if not os.path.exists(ruta_local) and not self.dropbox.descargar_archivo(ruta_dropbox, ruta_local):
            print(f"✨ [ExcelService] Archivo '{nombre_archivo}' no hallado en Dropbox. Inicializando nuevo mes...")
            ruta_origen_base = os.path.join(self.base_dir, plantilla_base_nombre)
            
            if not os.path.exists(ruta_origen_base):
                raise FileNotFoundError(f"❌ Falta la plantilla macro crítica en recursos locales: {ruta_origen_base}")
            
            # Copiamos la plantilla virgen a la zona temporal
            shutil.copy(ruta_origen_base, ruta_local)
            
            # Abrimos para asegurar la pestaña objetivo
            wb = openpyxl.load_workbook(ruta_local, keep_vba=True, data_only=False)
            self._asegurar_pestana_desde_plantilla(wb, pestana_target, fecha_str)
            wb.save(ruta_local)
            wb.close()
            time.sleep(1.5)
            # Subimos el archivo inicializado

            self.dropbox.subir_archivo(ruta_local, ruta_dropbox)

        # 2. Cargar el libro garantizado
        wb = openpyxl.load_workbook(ruta_local, keep_vba=True, data_only=False)
        self._asegurar_pestana_desde_plantilla(wb, pestana_target, fecha_str)
            
        return wb

    # ========================================================
    #       🌅 INYECCIÓN MATUTINA EN DOBLE PESTAÑA
    # ========================================================

    def _calcular_fecha_previa(self, fecha_str: str) -> str:
        """Si es Lunes, el día previo operativo fue Viernes (-3 días). Si no, -1 día."""
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        dias_atras = 3 if dt.weekday() == 0 else 1
        return (dt - timedelta(days=dias_atras)).strftime("%Y-%m-%d")

    def _escribir_promesa_dia_previo(self, ws, col: str, metas: dict):
        """Escribe en la hoja del Día Previo en las filas reales: 24 (UDVD), 25 (Visitas) y 26 (CxC)"""
        if "meta_udvd" in metas:
            ws[f"{col}{ContactoMatutinoMap.FILA_PREVIO_UDVD}"] = float(metas.get("meta_udvd", 0.0))
        if "meta_activaciones" in metas:
            ws[f"{col}{ContactoMatutinoMap.FILA_PREVIO_VISITAS}"] = int(metas.get("meta_activaciones", 0))
        if "meta_cxc" in metas:
            ws[f"{col}{ContactoMatutinoMap.FILA_PREVIO_CXC}"] = float(metas.get("meta_cxc", 0.0))
        print(f"  📌 Promesa inyectada en Día Previo ({ws.title}) -> Col {col} | Filas {ContactoMatutinoMap.FILA_PREVIO_UDVD}, {ContactoMatutinoMap.FILA_PREVIO_VISITAS}, {ContactoMatutinoMap.FILA_PREVIO_CXC}")

    def _escribir_meta_dia_actual(self, ws, col: str, metas: dict):
        """Escribe en la hoja del Día Actual en las filas reales: 11 (UDVD), 14 (Visitas) y 17 (CxC)"""
        if "meta_udvd" in metas:
            ws[f"{col}{ContactoMatutinoMap.FILA_META_UDVD}"] = float(metas.get("meta_udvd", 0.0))
        if "meta_activaciones" in metas:
            ws[f"{col}{ContactoMatutinoMap.FILA_META_VISITAS}"] = int(metas.get("meta_activaciones", 0))
        if "meta_cxc" in metas:
            ws[f"{col}{ContactoMatutinoMap.FILA_META_CXC}"] = float(metas.get("meta_cxc", 0.0))
        print(f"  🎯 Meta del Día inyectada en Día Actual ({ws.title}) -> Col {col} | Filas {ContactoMatutinoMap.FILA_META_UDVD}, {ContactoMatutinoMap.FILA_META_VISITAS}, {ContactoMatutinoMap.FILA_META_CXC}")

    def inyectar_plan_matutino_doble(self, ruta: int, fecha_target: str, metas: dict) -> bool:
        """
        [FUNCIÓN MAESTRA MATUTINA DEFINITIVA]
        Garantiza el libro, clona desde 'PLANTILLA' las pestañas previa y actual si no existen,
        inyecta las metas y realiza una sola subida limpia a Dropbox.
        """
        archivo_matutino, _, pestana_actual = self._calcular_nombres(fecha_target)
        fecha_previa = self._calcular_fecha_previa(fecha_target)
        _, _, pestana_previa = self._calcular_nombres(fecha_previa)

        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        mapeo_columnas = ContactoMatutinoMap.COLUMNAS_RUTAS
        col = mapeo_columnas.get(ruta)

        if not col:
            print(f"⚠️ [ExcelService] La Ruta {ruta} no tiene columna asignada en el mapeo matutino.")
            return False

        try:
            # 1. Garantizamos el archivo en local
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target="PLANTILLA",
                ruta_local=ruta_local
            )

            # 2. Aseguramos pestaña PREVIA e inyectamos
            ws_previo = self._asegurar_pestana_desde_plantilla(wb, pestana_previa, fecha_previa)
            self._escribir_promesa_dia_previo(ws_previo, col, metas)

            # 3. Aseguramos pestaña ACTUAL e inyectamos
            ws_actual = self._asegurar_pestana_desde_plantilla(wb, pestana_actual, fecha_target)
            self._escribir_meta_dia_actual(ws_actual, col, metas)

            # 4. Guardamos localmente y subimos a Dropbox
            wb.save(ruta_local)
            wb.close()
            time.sleep(1.5)
            print(f"📤 [ExcelService] Sincronizando inyección matutina doble a Dropbox -> {ruta_dropbox}")
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)

        except Exception as e:
            print(f"❌ [ExcelService] Error en inyección matutina doble: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)

    # ========================================================
    #       🌙 INYECCIÓN DE CIERRE NOCTURNO Y COBRANZA
    # ========================================================

    def inyectar_cierre_nocturno_excel(self, ruta: int, fecha: str, datos_cierre: dict) -> bool:
        """
        [REGISTRO ESTÁNDAR NOCTURNO]
        Inyecta en la pestaña de la fecha especificada:
        - Fila 12: UDVD (Unidades Logradas)
        - Fila 15: Visitas Efectivas
        - Fila 18: CXC Lograda ($)
        """
        archivo_matutino, _, pestana_target = self._calcular_nombres(fecha)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        mapeo_columnas = ContactoMatutinoMap.COLUMNAS_RUTAS
        col = mapeo_columnas.get(ruta)

        if not col:
            print(f"⚠️ [ExcelService] La Ruta {ruta} no tiene columna asignada en el cierre nocturno.")
            return False

        try:
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target=pestana_target,
                ruta_local=ruta_local,
                fecha_str=fecha
            )
            ws = wb[pestana_target]

            if "real_udvd" in datos_cierre:
                ws[f"{col}{ContactoMatutinoMap.FILA_REAL_UDVD}"] = float(datos_cierre.get("real_udvd", 0.0))
            if "real_activaciones" in datos_cierre:
                ws[f"{col}{ContactoMatutinoMap.FILA_REAL_VISITAS}"] = int(datos_cierre.get("real_activaciones", 0))
            if "real_cobranza" in datos_cierre:
                ws[f"{col}{ContactoMatutinoMap.FILA_REAL_CXC}"] = float(datos_cierre.get("real_cobranza", 0.0))

            print(f"🌙 [ExcelService] Cierre nocturno inyectado en {pestana_target} -> Col {col} | UDVD: Fila {ContactoMatutinoMap.FILA_REAL_UDVD}, Visitas: Fila {ContactoMatutinoMap.FILA_REAL_VISITAS}, CXC: Fila {ContactoMatutinoMap.FILA_REAL_CXC}")

            wb.save(ruta_local)
            wb.close()
            time.sleep(1.5)
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)
        except Exception as e:
            print(f"❌ [ExcelService] Error crítico en inyección nocturna: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)

    def actualizar_cobranza_acumulada_excel(self, lote_cobros: dict, fecha_str: str) -> bool:
        """
        [FUNCIÓN AUXILIAR DE RELEVO DE COBRANZA]
        lote_cobros: { 10: 1500.0, 15: 3200.0 }
        Escribe en la matriz P9:P20 según el día:
        - Lunes: Columna R | Martes: Columna S | Miércoles: Columna T | Jueves: Columna U
        - Viernes: SE IGNORA COMPLETAMENTE.
        """
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        dia_semana = dt.weekday()

        if dia_semana >= 4:
            print(f"ℹ️ [ExcelService] Fecha {fecha_str} es Viernes/Fin de semana. Se omite el relevo de cobranza.")
            return True

        mapa_columnas_dias = ContactoMatutinoMap.COLUMNAS_DIAS_COBRANZA
        col_dia = mapa_columnas_dias.get(dia_semana)

        archivo_matutino, _, pestana_target = self._calcular_nombres(fecha_str)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        try:
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target=pestana_target,
                ruta_local=ruta_local,
                fecha_str=fecha_str
            )
            ws = wb[pestana_target]

            print(f"💵 [ExcelService] Relevando cobranza para el día {ws.title} en Columna {col_dia}...")

            for ruta_id, monto in lote_cobros.items():
                if monto is None or float(monto) <= 0:
                    continue

                encontrado = False
                for fila in range(9, 21):
                    val_celda = ws[f"P{fila}"].value
                    if val_celda is not None and str(ruta_id) == str(val_celda).strip():
                        ws[f"{col_dia}{fila}"] = float(monto)
                        print(f"  🎯 Ruta {ruta_id} encontrada en P{fila} ➔ Cobranza de ${monto} asignada en {col_dia}{fila}")
                        encontrado = True
                        break

                if not encontrado:
                    print(f"⚠️ [ExcelService] La Ruta {ruta_id} no se encontró en el rango P9:P20.")

            wb.save(ruta_local)
            wb.close()
            time.sleep(1.5)
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)
        except Exception as e:
            print(f"❌ [ExcelService] Error en relevo de cobranza acumulada: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)

    # ========================================================
    #               🔍 MÉTODOS DE LECTURA AUTÓNOMOS
    # ========================================================

    def extraer_operacion_diaria_vendedor(self, ruta: int, fecha: str) -> dict:
        """[SR] Lee registros garantizando primero la existencia del archivo/pestaña."""
        archivo_matutino, _, pestana_target = self._calcular_nombres(fecha)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        
        try:
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target=pestana_target,
                ruta_local=ruta_local,
                fecha_str=fecha
            )
            wb.close()
            
            wb_lectura = openpyxl.load_workbook(ruta_local, data_only=True)
            ws = wb_lectura[pestana_target]
            
            mapeo_columnas = ContactoMatutinoMap.COLUMNAS_RUTAS
            col = mapeo_columnas.get(ruta)
            
            if col:
                data = {
                    "meta_udvd": ws[f"{col}{ContactoMatutinoMap.FILA_META_UDVD}"].value or 0.0,
                    "real_udvd": ws[f"{col}{ContactoMatutinoMap.FILA_REAL_UDVD}"].value or 0.0,
                    "meta_activaciones": ws[f"{col}{ContactoMatutinoMap.FILA_META_VISITAS}"].value or 0,
                    "real_activaciones": ws[f"{col}{ContactoMatutinoMap.FILA_REAL_VISITAS}"].value or 0,
                    "meta_cxc": ws[f"{col}{ContactoMatutinoMap.FILA_META_CXC}"].value or 0.0,
                    "real_cxc": ws[f"{col}{ContactoMatutinoMap.FILA_REAL_CXC}"].value or 0.0
                }
                wb_lectura.close()
                return data
            wb_lectura.close()
            return {}
        except Exception as e:
            print(f"🪵 [ExcelService] Error autónomo leyendo ventas: {e}")
            return {}
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)

    def extraer_caja_nocturna_dia(self, fecha: str) -> list:
        """[SR] Lee la matriz contable garantizando primero la existencia del reporte."""
        _, archivo_cobranza, pestana_target = self._calcular_nombres(fecha)
        ruta_local = os.path.join(self.tmp_dir, archivo_cobranza)
        
        try:
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_cobranza,
                plantilla_base_nombre="REPORTE_DIARIO_COBRANZA_BASE.xlsm",
                pestana_target=pestana_target,
                ruta_local=ruta_local,
                fecha_str=fecha
            )
            wb.close()
            
            df = pd.read_excel(ruta_local, sheet_name=pestana_target)
            return df.to_dict(orient="records")
        except Exception as e:
            print(f"🪵 [ExcelService] Error autónomo leyendo cobranzas: {e}")
            return []
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)

    # ========================================================
    #   📊 INYECCIÓN MAESTRA DE CUOTAS EN PESTAÑA CONTROL
    # ========================================================

    def actualizar_cuotas_control_excel(self, fecha_str: str, lote_cuotas: dict) -> bool:
        """Inyecta las cuotas presentes (>0) en la pestaña CONTROL"""
        archivo_matutino, _, _ = self._calcular_nombres(fecha_str)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        try:
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target="CONTROL",
                ruta_local=ruta_local
            )
            ws = wb[ContactoMatutinoMap.PESTANA_CONTROL]

            print(f"✏️ [ExcelService] Inyectando cuotas en pestaña CONTROL...")

            for ruta_id, metas in lote_cuotas.items():
                codigo_target = f"R-{ruta_id}"
                encontrado = False

                for fila in range(1, 51):
                    celda_o = ws[f"O{fila}"].value
                    celda_b = ws[f"B{fila}"].value
                
                    if celda_o == codigo_target or celda_b == codigo_target:
                        # SOLO actualiza si la métrica existe en el dict Y es mayor que 0
                        if "udvd" in metas and float(metas["udvd"]) > 0:
                            ws[f"R{fila}"] = float(metas["udvd"])
                        if "cobranza" in metas and float(metas["cobranza"]) > 0:
                            ws[f"S{fila}"] = float(metas["cobranza"])
                        if "visitas" in metas and float(metas["visitas"]) > 0:
                            ws[f"T{fila}"] = float(metas["visitas"])

                        print(f"  🎯 {codigo_target} en fila {fila} ➔ Modificado: {metas}")
                        encontrado = True
                        break
                
                if not encontrado:
                    print(f"⚠️ [ExcelService] No se encontró el código {codigo_target} en CONTROL.")

            wb.save(ruta_local)
            wb.close()
            print("⏳ [ExcelService] Esperando 1.5s para liberar el lock de Dropbox...")
            time.sleep(1.5)  # <--- ESTE SEGUNDO Y MEDIO SALVA LA VIDA
            
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)

        except Exception as e:
            print(f"❌ [ExcelService] Error crítico seteando cuotas en CONTROL: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)
    def resetear_libro_mensual_excel(self, fecha_str: str) -> bool:
        """Replica la macro 'EjecutarMantenimiento' reseteando CONTROL y purgando pestañas"""
        archivo_matutino, _, _ = self._calcular_nombres(fecha_str)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        HOJAS_PROTEGIDAS = [ContactoMatutinoMap.PESTANA_CONTROL, ContactoMatutinoMap.PESTANA_PLANTILLA]

        try:
            wb_init = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target="CONTROL",
                ruta_local=ruta_local
            )
            wb_init.close()

            with openpyxl.load_workbook(ruta_local, keep_vba=True, data_only=False) as wb:
                if "CONTROL" in wb.sheetnames:
                    ws_control = wb["CONTROL"]
                    ws_control["J4"] = None

                    for row in range(2, 51):
                        ws_control[f"R{row}"] = None
                        ws_control[f"S{row}"] = None
                        ws_control[f"T{row}"] = None

                    print("🧹 [ExcelService] Campos J4 y columnas de cuota reseteados en CONTROL.")

                hojas_a_borrar = [s for s in wb.sheetnames if s.upper() not in HOJAS_PROTEGIDAS]
                for sheet_name in hojas_a_borrar:
                    print(f"🗑️ [ExcelService] Purgando pestaña del mes: {sheet_name}")
                    del wb[sheet_name]

                wb.save(ruta_local)

            print(f"📤 [ExcelService] Subiendo libro reseteado a Dropbox -> {ruta_dropbox}")
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)

        except Exception as e:
            print(f"❌ [ExcelService] Error crítico durante el mantenimiento/reseteo: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)
    def limpiar_operacion_ruta_excel(self, ruta: int, fecha_str: str, modo: str = "COMPLETO") -> bool:
        """
        [LIMPIEZA FÍSICA EN EXCEL]
        Resetea a 0 las celdas específicas de una ruta sin borrar filas ni alterar fórmulas.
        """
        archivo_matutino, _, pestana_target = self._calcular_nombres(fecha_str)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        mapeo_columnas = ContactoMatutinoMap.COLUMNAS_RUTAS
        col = mapeo_columnas.get(ruta)

        if not col:
            print(f"⚠️ [ExcelService] La Ruta {ruta} no tiene columna asignada para limpiar.")
            return False

        try:
            wb = self._garantizar_archivo_y_pestana(
                nombre_archivo=archivo_matutino,
                plantilla_base_nombre="CONTACTO_MATUTINO_BASE.xlsm",
                pestana_target=pestana_target,
                ruta_local=ruta_local,
                fecha_str=fecha_str
            )
            ws = wb[pestana_target]

            modo_clean = modo.upper().strip()

            # --- LIMPIAR PLAN MATUTINO (Filas 11, 14, 17) ---
            if modo_clean in ["MATUTINO", "COMPLETO"]:
                ws[f"{col}{ContactoMatutinoMap.FILA_META_UDVD}"] = 0.0  # UDVD Meta
                ws[f"{col}{ContactoMatutinoMap.FILA_META_VISITAS}"] = 0    # Visitas Meta
                ws[f"{col}{ContactoMatutinoMap.FILA_META_CXC}"] = 0.0  # CxC Meta
                print(f"🧹 [ExcelService] Plan Matutino limpiado en Excel para Ruta {ruta} ({pestana_target})")

            # --- LIMPIAR CIERRE NOCTURNO (Filas 12, 15, 18) ---
            if modo_clean in ["NOCTURNO", "COMPLETO"]:
                ws[f"{col}{ContactoMatutinoMap.FILA_REAL_UDVD}"] = 0.0  # UDVD Real
                ws[f"{col}{ContactoMatutinoMap.FILA_REAL_VISITAS}"] = 0    # Visitas Real
                ws[f"{col}{ContactoMatutinoMap.FILA_REAL_CXC}"] = 0.0  # CxC Real
                print(f"🧹 [ExcelService] Cierre Nocturno limpiado en Excel para Ruta {ruta} ({pestana_target})")

            wb.save(ruta_local)
            wb.close()
            time.sleep(1.5)
            # Sincronizamos el Excel limpio con Dropbox
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)

        except Exception as e:
            print(f"❌ [ExcelService] Error limpiando celdas en Excel: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)
                
    # ========================================================
    #       🔄 MÓDULO JOB NOCTURNO: ESCÁNER PASIVO
    # ========================================================

    def escanear_libro_mensual_completo(self, fecha_str: str) -> dict:
        """
        [ESCÁNER PASIVO - EL EXCEL MANDA]
        Descarga el libro mensual de Dropbox si existe. Si NO existe, aborta.
        Lee cuotas de CONTROL y la operación real de cada pestaña diaria.
        NO edita ni guarda nada en Excel.
        """
        archivo_matutino, _, _ = self._calcular_nombres(fecha_str)
        ruta_local = os.path.join(self.tmp_dir, archivo_matutino)
        ruta_dropbox = f"{self.MAIN_FOLDER}{archivo_matutino}"

        # 1. ¿Dónde estoy? Si el archivo NO existe en Dropbox, abortamos en silencio
        if not self.dropbox.descargar_archivo(ruta_dropbox, ruta_local):
            print(f"ℹ️ [ExcelService] El archivo '{archivo_matutino}' no existe en Dropbox. Sincronización omitida.")
            return {}

        resumen_escaneo = {
            "cuotas_control": {},
            "operaciones_diarias": {}
        }

        try:
            # Cargar con data_only=True para leer los valores finales calculados por formulas/macros
            wb = openpyxl.load_workbook(ruta_local, data_only=True)
            
            # ----------------------------------------------------
            # FASE 1: ESCANEO DE CUOTAS EN PESTAÑA 'CONTROL'
            # ----------------------------------------------------
            pestana_control = None
            for s in wb.sheetnames:
                if s.strip().upper() == "CONTROL":
                    pestana_control = wb[s]
                    break

            if pestana_control:
                print("🔍 [Escáner] Procesando cuotas en pestaña CONTROL (R13:T20)...")
                for fila in range(13, 21):
                    val_ruta_raw = pestana_control[f"O{fila}"].value
                    if not val_ruta_raw:
                        continue

                    # Extraer el número de ruta de 'R-10', 'R-15', etc.
                    val_str = str(val_ruta_raw).strip().upper()
                    if not val_str.startswith("R-"):
                        continue
                    
                    try:
                        ruta_id = int(val_str.replace("R-", ""))
                    except ValueError:
                        continue

                    u_val = pestana_control[f"R{fila}"].value
                    c_val = pestana_control[f"S{fila}"].value
                    v_val = pestana_control[f"T{fila}"].value

                    metas = {}
                    # Regla estricta: Solo si es numérico y mayor a 0
                    if isinstance(u_val, (int, float)) and u_val > 0:
                        metas["udvd"] = float(u_val)
                    if isinstance(c_val, (int, float)) and c_val > 0:
                        metas["cobranza"] = float(c_val)
                    if isinstance(v_val, (int, float)) and v_val > 0:
                        metas["visitas"] = float(v_val)

                    if metas:
                        resumen_escaneo["cuotas_control"][ruta_id] = metas

            # ----------------------------------------------------
            # FASE 2: ESCANEO DE PESTAÑAS DIARIAS (IGNORA CONTROL Y PLANTILLA)
            # ----------------------------------------------------
            HOJAS_IGNORADAS = ["CONTROL", "PLANTILLA"]
            mapeo_columnas = ContactoMatutinoMap.COLUMNAS_RUTAS

            pestañas_operativas = [
                s for s in wb.sheetnames 
                if s.strip().upper() not in HOJAS_IGNORADAS
            ]

            print(f"🔍 [Escáner] Procesando {len(pestañas_operativas)} pestañas diarias...")

            for sheet_name in pestañas_operativas:
                ws = wb[sheet_name]

                # Intentar parsear fecha desde la celda G4 o usar el nombre de la pestaña
                fecha_pestana = None
                g4_val = ws[ContactoMatutinoMap.CELDA_FECHA_G4].value
                if g4_val:
                    try:
                        if isinstance(g4_val, datetime):
                            fecha_pestana = g4_val.strftime("%Y-%m-%d")
                        else:
                            dt = datetime.strptime(str(g4_val).strip(), "%d/%m/%Y")
                            fecha_pestana = dt.strftime("%Y-%m-%d")
                    except ValueError:
                        pass

                if not fecha_pestana:
                    continue

                if fecha_pestana not in resumen_escaneo["operaciones_diarias"]:
                    resumen_escaneo["operaciones_diarias"][fecha_pestana] = {}

                # Extraer datos por columna de ruta
                for ruta_id, col in mapeo_columnas.items():
                    m_u = ws[f"{col}{ContactoMatutinoMap.FILA_META_UDVD}"].value or 0.0
                    r_u = ws[f"{col}{ContactoMatutinoMap.FILA_REAL_UDVD}"].value or 0.0
                    m_v = ws[f"{col}{ContactoMatutinoMap.FILA_META_VISITAS}"].value or 0
                    r_v = ws[f"{col}{ContactoMatutinoMap.FILA_REAL_VISITAS}"].value or 0
                    m_c = ws[f"{col}{ContactoMatutinoMap.FILA_META_CXC}"].value or 0.0
                    r_c = ws[f"{col}{ContactoMatutinoMap.FILA_REAL_CXC}"].value or 0.0

                    # Si hay algún dato cargado en las celdas
                    if any([m_u, r_u, m_v, r_v, m_c, r_c]):
                        resumen_escaneo["operaciones_diarias"][fecha_pestana][ruta_id] = {
                            "meta_udvd": float(m_u if isinstance(m_u, (int, float)) else 0.0),
                            "real_udvd": float(r_u if isinstance(r_u, (int, float)) else 0.0),
                            "meta_activaciones": int(m_v if isinstance(m_v, (int, float)) else 0),
                            "real_activaciones": int(r_v if isinstance(r_v, (int, float)) else 0),
                            "meta_cxc": float(m_c if isinstance(m_c, (int, float)) else 0.0),
                            "real_cxc": float(r_c if isinstance(r_c, (int, float)) else 0.0)
                        }

            wb.close()
            return resumen_escaneo

        except Exception as e:
            print(f"❌ [ExcelService] Error durante el escaneo pasivo del libro: {e}")
            return {}
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)
                
# Método auxiliar de cálculo de semanas y construcción del libro dentro de ExcelService

    def _obtener_nombre_mes_capitalizado(self, mes_num: int) -> str:
        meses = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        return meses.get(mes_num, "Enero")

    def _calcular_semanas_del_mes(self, anio: int, mes: int) -> list:
        """
        Calcula los rangos de días hábiles (Lunes a Viernes) para cada semana del mes.
        Retorna lista de listas de objetos datetime.date de 5 días por semana.
        """
        import calendar
        cal = calendar.Calendar(firstweekday=0) # Lunes como primer día
        dias_mes = cal.monthdatescalendar(anio, mes)
        
        semanas_validas = []
        for semana in dias_mes:
            # Tomamos solo de Lunes (0) a Viernes (4)
            dias_laborables = semana[:5]
            # Si al menos un día laborable cae dentro del mes objetivo, cuenta como semana del mes
            if any(d.month == mes for d in dias_laborables):
                semanas_validas.append(dias_laborables)
                
        return semanas_validas

    def _garantizar_archivo_cobranza_mensual(self, fecha_str: str) -> str:
        """
        Verifica la existencia del archivo de cobranza en Dropbox.
        Si no existe, clona la plantilla base, genera N pestañas de semanas (Semana I, II...)
        e inyecta los textos de encabezados 'COBRANZA DEL LUNES DD/MM/YYYY'.
        """
        from config.excel_map_config import CobranzaDiariaMap as MapCD
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        
        nombre_mes = self._obtener_nombre_mes_capitalizado(dt.month)
        anio_dos_digitos = str(dt.year)[2:]
        nombre_archivo = f"REPORTE DIARIO DE COBRANZA {nombre_mes}-{anio_dos_digitos}.xlsx"
        
        ruta_local = os.path.join(self.tmp_dir, nombre_archivo)
        ruta_dropbox = f"{MapCD.MAIN_FOLDER}{nombre_archivo}"

        # 1. Intentar descargar de Dropbox
        if os.path.exists(ruta_local) or self.dropbox.descargar_archivo(ruta_dropbox, ruta_local):
            return nombre_archivo

        # 2. Si NO existe en Dropbox, construir desde la plantilla base
        ruta_base = os.path.join(self.base_dir, MapCD.PLANTILLA_BASE_NOMBRE)
        if not os.path.exists(ruta_base):
            raise FileNotFoundError(f"❌ Falta la plantilla base de cobranza: {ruta_base}")

        shutil.copy(ruta_base, ruta_local)
        wb = openpyxl.load_workbook(ruta_local, keep_vba=True, data_only=False)

        # Determinar la hoja plantilla inicial
        sheet_base = wb.worksheets[0]

        semanas = self._calcular_semanas_del_mes(dt.year, dt.month)
        romanos = ["I", "II", "III", "IV", "V", "VI"]

        hojas_creadas = []

        for idx, semana_dias in enumerate(semanas):
            nombre_hoja = f"Semana {romanos[idx]}"
            ws_semana = wb.copy_worksheet(sheet_base)
            ws_semana.title = nombre_hoja
            hojas_creadas.append(ws_semana)

            # Inyectar las 5 fechas en los encabezados
            for i_dia, dia_dt in enumerate(semana_dias):
                cfg_dia = MapCD.CONFIG_DIAS.get(i_dia)
                if cfg_dia:
                    celda_encabezado = cfg_dia["encabezado"]
                    nombre_dia_txt = MapCD.NOMBRES_DIAS_TEXTO.get(i_dia, "LUNES")
                    fecha_formateada = dia_dt.strftime("%d/%m/%Y")
                    
                    # Formato requerido: "COBRANZA DEL LUNES 03/08/2026"
                    texto_completo = f"COBRANZA DEL {nombre_dia_txt} {fecha_formateada}"
                    ws_semana[celda_encabezado] = texto_completo

        # Eliminar la pestaña base original
        wb.remove(sheet_base)

        wb.save(ruta_local)
        wb.close()

        # Sincronizar archivo nuevo inicializado con Dropbox
        self.dropbox.subir_archivo(ruta_local, ruta_dropbox)
        return nombre_archivo

    def inyectar_cobranza_diaria_excel(self, ruta: int, fecha_str: str, efectivo: float, zelle: float, bs: float) -> bool:
        """
        Inyecta la cobranza desglosada (Efectivo, Zelle, Bolívares) en el día
        y semana correspondientes dentro del Reporte Diario de Cobranza.
        """
        
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        dia_semana = dt.weekday()
        
        if dia_semana >= 5: # Ignorar fines de semana (Sábado/Domingo)
            print(f"ℹ️ [ExcelService] {fecha_str} es fin de semana. No aplica reporte de cobranza diario.")
            return True

        col = MapCD.COLUMNAS_RUTAS.get(ruta)
        if not col:
            print(f"⚠️ [ExcelService] La Ruta {ruta} no tiene columna asignada en CobranzaDiariaMap.")
            return False

        try:
            nombre_archivo = self._garantizar_archivo_cobranza_mensual(fecha_str)
            ruta_local = os.path.join(self.tmp_dir, nombre_archivo)
            ruta_dropbox = f"{MapCD.MAIN_FOLDER}{nombre_archivo}"

            wb = openpyxl.load_workbook(ruta_local, keep_vba=True, data_only=False)

            # Ubicar la semana correspondiente a la fecha
            semanas = self._calcular_semanas_del_mes(dt.year, dt.month)
            romanos = ["I", "II", "III", "IV", "V", "VI"]
            
            idx_semana_target = None
            for idx, sem_dias in enumerate(semanas):
                if dt.date() in sem_dias:
                    idx_semana_target = idx
                    break

            if idx_semana_target is None:
                # Fallback por si la fecha cae en solapamiento de mes
                idx_semana_target = 0

            nombre_hoja_target = f"Semana {romanos[idx_semana_target]}"
            if nombre_hoja_target not in wb.sheetnames:
                ws = wb.worksheets[0]
            else:
                ws = wb[nombre_hoja_target]

            cfg_dia = MapCD.CONFIG_DIAS.get(dia_semana)

            # Inyección de los valores por divisa
            ws[f"{col}{cfg_dia['efectivo']}"] = float(efectivo or 0.0)
            ws[f"{col}{cfg_dia['zelle']}"] = float(zelle or 0.0)
            ws[f"{col}{cfg_dia['bs']}"] = float(bs or 0.0)

            wb.save(ruta_local)
            wb.close()
            time.sleep(1.5)
            print(f"💵 [ExcelService] Cobranza inyectada en '{nombre_archivo}' -> {nombre_hoja_target} ({fecha_str}) Ruta {ruta}")
            return self.dropbox.subir_archivo(ruta_local, ruta_dropbox)

        except Exception as e:
            print(f"❌ [ExcelService] Error inyectando cobranza diaria: {e}")
            return False
        finally:
            if os.path.exists(ruta_local):
                os.remove(ruta_local)